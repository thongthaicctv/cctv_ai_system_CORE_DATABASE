import os
import sqlite3

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from core.resource_paths import ensure_app_file
from hr.report_db import query_report_by_date, query_report_by_employee
from hr.employee_manager import load_employees

FONT_NAME = "Times New Roman"
FONT_SIZE = 13

HEADER_FILL = "1F4E78"
HEADER_FONT = "FFFFFF"
ALT_FILL = "F2F2F2"
BORDER_COLOR = "000000"


def _style_report(ws, title_text: str, last_col: int, last_row: int):
    thin = Side(style="thin", color=BORDER_COLOR)

    ws.merge_cells(
        start_row=1,
        start_column=1,
        end_row=1,
        end_column=last_col
    )

    title = ws.cell(1, 1)
    title.value = title_text
    title.font = Font(
        name=FONT_NAME,
        size=16,
        bold=True
    )
    title.alignment = Alignment(
        horizontal="center",
        vertical="center"
    )

    ws.row_dimensions[1].height = 32
    ws.row_dimensions[2].height = 26

    for cell in ws[2]:
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.font = Font(
            name=FONT_NAME,
            size=FONT_SIZE,
            bold=True,
            color=HEADER_FONT
        )
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True
        )
        cell.border = Border(
            top=thin,
            bottom=thin,
            left=thin,
            right=thin
        )

    for row_idx in range(3, last_row + 1):
        ws.row_dimensions[row_idx].height = 22

        for col_idx in range(1, last_col + 1):
            cell = ws.cell(row_idx, col_idx)

            cell.font = Font(
                name=FONT_NAME,
                size=FONT_SIZE
            )

            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
                wrap_text=True
            )

            cell.border = Border(
                top=thin,
                bottom=thin,
                left=thin,
                right=thin
            )

            if row_idx % 2 == 0:
                cell.fill = PatternFill("solid", fgColor=ALT_FILL)

    ws.freeze_panes = "A3"

    if last_row >= 2:
        ws.auto_filter.ref = f"A2:{get_column_letter(last_col)}{last_row}"

    for col_idx in range(1, last_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0

        for row_idx in range(1, last_row + 1):
            value = ws.cell(row_idx, col_idx).value
            max_len = max(max_len, len(str(value or "")))

        ws.column_dimensions[col_letter].width = min(max_len + 4, 28)


def export_total_excel(from_date: str, to_date: str, save_path: str):
    rows = query_report_by_date(from_date, to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao tong hop"

    headers = [
        "Ngày",
        "Mã đơn",
        "Mã NV",
        "Tên nhân viên",
        "Bộ phận",
        "Camera",
        "Thời gian bắt đầu",
        "Thời gian kết thúc",
        "Độ dài (giây)",
        "Dung lượng (MB)",
    ]

    ws.append(["BÁO CÁO TỔNG HỢP"])
    ws.append(headers)

    for r in rows:
        ws.append([
            r.get("date", ""),
            r.get("order_code", ""),
            r.get("employee_id", ""),
            r.get("employee_name", ""),
            r.get("department", ""),
            r.get("camera_name", ""),
            r.get("start_time", ""),
            r.get("end_time", ""),
            r.get("duration_sec", 0),
            r.get("file_size_mb", 0),
        ])

    last_row = ws.max_row
    last_col = len(headers)

    _style_report(ws, "BÁO CÁO TỔNG HỢP", last_col, last_row)

    wb.save(save_path)
    return save_path


def export_employee_excel(from_date: str, to_date: str, save_path: str):
    rows = query_report_by_employee(from_date, to_date)

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao nhan vien"

    headers = [
        "Mã NV",
        "Tên nhân viên",
        "Bộ phận",
        "Chức vụ",
        "Tổng đơn",
        "Tổng video",
        "Tổng thời lượng (giây)",
        "Tổng thời lượng (phút)",
        "Tổng dung lượng (MB)",
    ]

    ws.append(["BÁO CÁO THEO NHÂN VIÊN"])
    ws.append(headers)

    for r in rows:
        duration_sec = r.get("total_duration_sec") or 0

        ws.append([
            r.get("employee_id", ""),
            r.get("employee_name", ""),
            r.get("department", ""),
            r.get("position", ""),
            r.get("total_orders", 0),
            r.get("total_videos", 0),
            duration_sec,
            round(duration_sec / 60, 2),
            round(r.get("total_size_mb") or 0, 2),
        ])

    last_row = ws.max_row
    last_col = len(headers)

    _style_report(ws, "BÁO CÁO THEO NHÂN VIÊN", last_col, last_row)

    wb.save(save_path)
    return save_path

def export_all_reports_excel(from_date: str, to_date: str, save_path: str):
    """
    Xuất 1 file Excel gồm 2 sheet:
    Sheet 1: Báo cáo tổng hợp
    Sheet 2: Báo cáo theo nhân viên
    Giữ nguyên font, cỡ chữ, căn giữa, height, filter theo form hiện tại.
    """

    wb = Workbook()

    # =========================
    # SHEET 1 - BÁO CÁO TỔNG HỢP
    # =========================
    rows_total = query_report_by_date(from_date, to_date)

    ws = wb.active
    ws.title = "Bao cao tong hop"

    headers_total = [
        "Ngày",
        "Mã đơn",
        "Mã NV",
        "Tên nhân viên",
        "Bộ phận",
        "Camera",
        "Thời gian bắt đầu",
        "Thời gian kết thúc",
        "Độ dài (giây)",
        "Dung lượng (MB)",
    ]

    ws.append(["BÁO CÁO TỔNG HỢP"])
    ws.append(headers_total)

    for r in rows_total:
        ws.append([
            r.get("date", ""),
            r.get("order_code", ""),
            r.get("employee_id", ""),
            r.get("employee_name", ""),
            r.get("department", ""),
            r.get("camera_name", ""),
            r.get("start_time", ""),
            r.get("end_time", ""),
            r.get("duration_sec", 0),
            r.get("file_size_mb", 0),
        ])

    _style_report(
        ws,
        "BÁO CÁO TỔNG HỢP",
        len(headers_total),
        ws.max_row
    )

    # =========================
    # SHEET 2 - BÁO CÁO THEO NHÂN VIÊN
    # =========================
    rows_emp = query_report_by_employee(from_date, to_date)

    ws2 = wb.create_sheet("Bao cao nhan vien")

    headers_emp = [
        "Mã NV",
        "Tên nhân viên",
        "Bộ phận",
        "Chức vụ",
        "Tổng đơn",
        "Tổng video",
        "Tổng thời lượng (giây)",
        "Tổng thời lượng (phút)",
        "Tổng dung lượng (MB)",
    ]

    ws2.append(["BÁO CÁO THEO NHÂN VIÊN"])
    ws2.append(headers_emp)

    for r in rows_emp:
        duration_sec = r.get("total_duration_sec") or 0

        ws2.append([
            r.get("employee_id", ""),
            r.get("employee_name", ""),
            r.get("department", ""),
            r.get("position", ""),
            r.get("total_orders", 0),
            r.get("total_videos", 0),
            duration_sec,
            round(duration_sec / 60, 2),
            round(r.get("total_size_mb") or 0, 2),
        ])

    _style_report(
        ws2,
        "BÁO CÁO THEO NHÂN VIÊN",
        len(headers_emp),
        ws2.max_row
    )

    wb.save(save_path)
    return save_path


def _handover_db_path():
    """
    DB nghiệp vụ đóng/giao hàng.
    Ưu tiên lấy theo thư mục gốc project:
    D:/PYTHON-TCR/cctv_ai_system_ffmpeg/db/packing.db
    """
    return ensure_app_file("db", "packing.db")

def _employee_name_map():
    """
    Map mã nhân viên -> thông tin nhân viên từ employees.json.
    """
    try:
        employees = load_employees()
    except Exception:
        employees = []

    data = {}

    for emp in employees:
        emp_id = str(emp.get("id", "")).strip()
        if not emp_id:
            continue

        data[emp_id.lower()] = {
            "id": emp_id,
            "name": str(emp.get("name", "")).strip(),
            "department": str(emp.get("department", "")).strip(),
            "position": str(emp.get("position", "")).strip(),
        }

    return data


def _employee_real_name(emp_code, emp_name="", emp_map=None):
    """
    Trả về tên thật theo mã nhân viên.
    Ưu tiên tên đã lưu trong DB, nếu trống thì dò employees.json.
    """
    emp_code = str(emp_code or "").strip()
    emp_name = str(emp_name or "").strip()

    if emp_name:
        return emp_name

    if emp_code and emp_map:
        found = emp_map.get(emp_code.lower())
        if found:
            return found.get("name", "")

    return ""

def export_handover_excel(from_date: str, to_date: str, save_path: str):
    """
    Xuất báo cáo giao hàng theo mẫu:

    Sheet 1: Báo cáo giao hàng tổng hợp
    Sheet 2: Chi tiết kiện nhỏ trong từng đơn lớn
    """

    db_path = _handover_db_path()
    emp_map = _employee_name_map()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao giao hang"

    headers = [
        "STT",
        "Mã đơn (thùng)",
        "Thời gian đóng",
        "Mã NV đóng",
        "Tên NV đóng",
        "Mã NV giao",
        "Tên NV giao",
        "Scanner đóng",
        "số lượng kiện nhỏ",
        "Chi tiết kiện nhỏ",
        "Kết quả",
        "Lỗi/Ghi chú",
        "Lần quét giao",
        "Lần quét thùng",
    ]

    ws.append(["BÁO CÁO GIAO HÀNG"])
    ws.append(headers)

    if not os.path.exists(db_path):
        ws.append([
            1,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Không tìm thấy DB",
            f"Không thấy file: {db_path}",
            "",
            "",
        ])

        _style_report(
            ws,
            "BÁO CÁO GIAO HÀNG",
            len(headers),
            ws.max_row
        )

        wb.save(save_path)
        return save_path

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    # Kiểm tra cột nhân viên giao nếu đã bổ sung vào DB
    try:
        cur.execute("PRAGMA table_info(handover_sessions)")
        handover_cols = {r["name"] for r in cur.fetchall()}
    except Exception:
        handover_cols = set()

    delivery_emp_code_select = ""
    delivery_emp_name_select = ""

    if "delivery_employee_code" in handover_cols:
        delivery_emp_code_select = ", h.delivery_employee_code AS delivery_employee_code"

    if "delivery_employee_name" in handover_cols:
        delivery_emp_name_select = ", h.delivery_employee_name AS delivery_employee_name"

    where = []
    params = []

    if from_date:
        where.append("date(h.created_at) >= ?")
        params.append(from_date)

    if to_date:
        where.append("date(h.created_at) <= ?")
        params.append(to_date)

    where_sql = ""
    if where:
        where_sql = "WHERE " + " AND ".join(where)

    cur.execute(f"""
        SELECT
            h.id,
            h.delivery_order_code,
            h.packing_order_code,
            h.packing_session_id,
            h.packing_scanner_id,
            h.delivery_scanner_id,
            h.result,
            h.error_message,
            h.first_scan_time,
            h.second_scan_time,
            h.created_at,

            p.master_order_code,
            p.employee_code AS packing_employee_code,
            p.employee_name AS packing_employee_name,
            p.start_time AS packing_start_time,
            p.end_time AS packing_end_time,
            
            p.total_items AS packing_total_items
            {delivery_emp_code_select}
            {delivery_emp_name_select}

        FROM handover_sessions h
        LEFT JOIN packing_sessions p
            ON h.packing_session_id = p.id

        {where_sql}

        ORDER BY h.created_at DESC, h.id DESC
    """, params)

    rows = cur.fetchall()

    # =========================================================
    # LẤY CHI TIẾT KIỆN NHỎ TỪ packing_items
    # =========================================================
    session_ids = []
    for r in rows:
        sid = r["packing_session_id"]
        if sid and sid not in session_ids:
            session_ids.append(sid)

    items_by_session = {}

    if session_ids:
        placeholders = ",".join(["?"] * len(session_ids))

        try:
            cur.execute("PRAGMA table_info(packing_items)")
            packing_item_cols = {r["name"] for r in cur.fetchall()}
        except Exception:
            packing_item_cols = set()

        delete_filter = ""
        if "is_deleted" in packing_item_cols:
            delete_filter = "AND IFNULL(is_deleted, 0) = 0"

        cur.execute(f"""
            SELECT
                id,
                session_id,
                master_order_code,
                item_code,
                scan_time,
                scan_index
            FROM packing_items
            WHERE session_id IN ({placeholders})
            {delete_filter}
            ORDER BY session_id ASC, scan_index ASC, id ASC
        """, session_ids)

        item_rows = cur.fetchall()

        for item in item_rows:
            sid = item["session_id"]
            items_by_session.setdefault(sid, []).append(item)

    conn.close()

    def result_text(value):
        value = str(value or "").strip()

        if value == "success":
            return "Giao vận chuyển thành công"

        if value == "failed_wrong_box":
            return "Sai thùng hàng"

        if value == "failed_not_packed":
            return "Đơn chưa được đóng"

        if value == "failed_already_delivered":
            return "Đơn đã giao trước đó"

        if value == "failed_missing_box_code":
            return "Chưa quét mã thùng"

        if value == "blocked_already_delivered":
            return "Chặn do đơn đã giao"

        return value

    # =========================================================
    # SHEET 1 - BÁO CÁO GIAO HÀNG TỔNG HỢP
    # =========================================================
    if not rows:
        ws.append([
            1,
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "Không có dữ liệu",
            "Không có dữ liệu giao hàng trong khoảng ngày đã chọn",
            "",
            "",
        ])
    else:
        for idx, r in enumerate(rows, start=1):
            session_id = r["packing_session_id"]

            ma_don_thung = (
                r["packing_order_code"]
                or r["master_order_code"]
                or r["delivery_order_code"]
                or ""
            )

            thoi_gian_dong = (
                r["packing_end_time"]
                or r["packing_start_time"]
                or ""
            )

            ma_nv_dong = r["packing_employee_code"] or ""
            ten_nv_dong = _employee_real_name(
                ma_nv_dong,
                r["packing_employee_name"],
                emp_map
            )

            delivery_employee_code = ""
            delivery_employee_name = ""

            try:
                row_keys = r.keys()

                if "delivery_employee_code" in row_keys:
                    delivery_employee_code = r["delivery_employee_code"] or ""

                if "delivery_employee_name" in row_keys:
                    delivery_employee_name = r["delivery_employee_name"] or ""

            except Exception:
                pass

            # Nếu DB cũ chưa có cột nhân viên giao thì fallback lấy scanner giao làm mã tạm
            ma_nv_giao = delivery_employee_code or ""
            ten_nv_giao = _employee_real_name(
                ma_nv_giao,
                delivery_employee_name,
                emp_map
            )


            scanner_dong = (
                r["packing_scanner_id"]
                or ""
            )

            item_rows = items_by_session.get(session_id, [])

            chi_tiet_kien_nho = "\n".join(
                [
                    f"{item['scan_index']}. {item['item_code']}"
                    for item in item_rows
                ]
            )

            so_luong_kien_nho = (
                len(item_rows)
                if item_rows
                else (r["packing_total_items"] or 0)
            )

            ws.append([
                idx,
                ma_don_thung,
                thoi_gian_dong,
                ma_nv_dong,
                ten_nv_dong,
                ma_nv_giao,
                ten_nv_giao,
                scanner_dong,
                so_luong_kien_nho,
                chi_tiet_kien_nho,
                result_text(r["result"]),
                r["error_message"] or "",
                r["first_scan_time"] or "",
                r["second_scan_time"] or "",
            ])
    last_row = ws.max_row
    last_col = len(headers)

    _style_report(
        ws,
        "BÁO CÁO GIAO HÀNG",
        last_col,
        last_row
    )

    # Tô màu cột Kết quả
    success_fill = PatternFill("solid", fgColor="C6EFCE")
    error_fill = PatternFill("solid", fgColor="FFC7CE")
    warning_fill = PatternFill("solid", fgColor="FFEB9C")

    for row_idx in range(3, last_row + 1):
        result_cell = ws.cell(row_idx, 11)
        text = str(result_cell.value or "")

        if text == "Giao vận chuyển thành công":
            result_cell.fill = success_fill
        elif text in ("Không có dữ liệu", "Không tìm thấy DB"):
            result_cell.fill = warning_fill
        elif text:
            result_cell.fill = error_fill

        # Chi tiết kiện nhỏ căn trái và wrap text
        ws.cell(row_idx, 10).alignment = Alignment(
            horizontal="left",
            vertical="top",
            wrap_text=True
        )

        # Lỗi/Ghi chú căn trái
        ws.cell(row_idx, 12).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        # Tăng chiều cao dòng nếu có nhiều kiện nhỏ
        detail_text = str(ws.cell(row_idx, 10).value or "")
        if "\n" in detail_text:
            line_count = detail_text.count("\n") + 1
            ws.row_dimensions[row_idx].height = min(22 + line_count * 12, 120)

    widths = {
        1: 8,    # STT
        2: 24,   # Mã đơn thùng
        3: 22,   # Thời gian đóng
        4: 16,   # Mã NV đóng
        5: 22,   # Tên NV đóng
        6: 16,   # Mã NV giao
        7: 22,   # Tên NV giao
        8: 18,   # Scanner đóng
        9: 20,   # số lượng kiện nhỏ
        10: 38,  # Chi tiết kiện nhỏ
        11: 24,  # Kết quả
        12: 42,  # Lỗi/Ghi chú
        13: 24,  # Lần quét giao
        14: 24,  # Lần quét thùng
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # =========================================================
    # SHEET 2 - CHI TIẾT KIỆN NHỎ
    # =========================================================
    ws_detail = wb.create_sheet("Chi tiet kien nho")

    headers_detail = [
        "STT",
        "Mã đơn (thùng)",
        "Packing Session ID",
        "Scanner đóng",
        "Mã NV đóng",
        "Tên NV đóng",
        "STT kiện nhỏ",
        "Mã kiện nhỏ",
        "Thời gian quét kiện",
        "Kết quả giao",
    ]

    ws_detail.append(["CHI TIẾT KIỆN NHỎ"])
    ws_detail.append(headers_detail)

    detail_index = 1

    for r in rows:
        session_id = r["packing_session_id"]
        item_rows = items_by_session.get(session_id, [])

        ma_don_thung = (
            r["packing_order_code"]
            or r["master_order_code"]
            or r["delivery_order_code"]
            or ""
        )

        ma_nv_dong = r["packing_employee_code"] or ""
        ten_nv_dong = _employee_real_name(
            ma_nv_dong,
            r["packing_employee_name"],
            emp_map
        )

        scanner_dong = r["packing_scanner_id"] or ""
        ket_qua_giao = result_text(r["result"])

        if not item_rows:
            ws_detail.append([
                detail_index,
                ma_don_thung,
                session_id or "",
                scanner_dong,
                ma_nv_dong,
                ten_nv_dong,
                "",
                "",
                "",
                ket_qua_giao,
            ])
            detail_index += 1
            continue

        for item in item_rows:
            ws_detail.append([
                detail_index,
                ma_don_thung,
                session_id or "",
                scanner_dong,
                ma_nv_dong,
                ten_nv_dong,
                item["scan_index"] or "",
                item["item_code"] or "",
                item["scan_time"] or "",
                ket_qua_giao,
            ])
            detail_index += 1

    _style_report(
        ws_detail,
        "CHI TIẾT KIỆN NHỎ",
        len(headers_detail),
        ws_detail.max_row
    )

    detail_widths = {
        1: 8,
        2: 24,
        3: 18,
        4: 16,
        5: 16,
        6: 22,
        7: 14,
        8: 32,
        9: 22,
        10: 22,
    }

    for col_idx, width in detail_widths.items():
        ws_detail.column_dimensions[get_column_letter(col_idx)].width = width

    for row_idx in range(3, ws_detail.max_row + 1):
        ws_detail.cell(row_idx, 8).alignment = Alignment(
            horizontal="left",
            vertical="center",
            wrap_text=True
        )

        result_cell = ws_detail.cell(row_idx, 10)
        text = str(result_cell.value or "")

        if text == "Giao vận chuyển thành công":
            result_cell.fill = success_fill
        elif text:
            result_cell.fill = error_fill

    wb.save(save_path)
    return save_path
