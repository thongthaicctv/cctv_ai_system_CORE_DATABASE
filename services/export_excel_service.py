# -*- coding: utf-8 -*-
"""Excel exports for the shared MySQL order database."""

from __future__ import annotations

from collections import defaultdict
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from services.mysql_client import MySQLClient


EMPLOYEE_HEADERS = [
    "STT",
    "Mã NV",
    "Họ & tên",
    "Bộ phận",
    "Điện thoại",
    "Trạng thái",
    "Ngày tạo",
    "Cập nhật",
]

ECOM_HEADERS = [
    "STT",
    "Mã đơn",
    "Sàn/Shop",
    "Khách hàng",
    "SĐT",
    "Trạng thái đơn",
    "Đóng hàng",
    "Băng truyền",
    "Vận chuyển",
    "Mã vận đơn",
    "Đơn vị VC",
    "Số video",
    "Video cuối",
    "Ghi chú",
    "Ngày tạo",
    "Cập nhật",
]

WHOLESALE_HEADERS = [
    "STT",
    "Mã đơn sỉ",
    "Khách hàng",
    "SĐT",
    "Trạng thái đơn",
    "Đóng hàng",
    "Chi tiết kiện nhỏ",
    "Vận chuyển",
    "Mã vận đơn",
    "Đơn vị VC",
    "Số video",
    "Video cuối",
    "Ghi chú",
    "Ngày tạo",
    "Cập nhật",
]


def _clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    return str(value)


def _parse_date(value: Optional[str]) -> Optional[str]:
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(value, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    raise ValueError(f"Ngày không hợp lệ: {value}. Dùng YYYY-MM-DD hoặc DD/MM/YYYY.")


def _where_by_date(from_date: Optional[str], to_date: Optional[str]) -> tuple[str, list[str]]:
    where = []
    params: list[str] = []
    from_date = _parse_date(from_date)
    to_date = _parse_date(to_date)

    if from_date:
        where.append("DATE(o.created_at) >= %s")
        params.append(from_date)
    if to_date:
        where.append("DATE(o.created_at) <= %s")
        params.append(to_date)

    return ("WHERE " + " AND ".join(where)) if where else "", params


def _table_exists(cur, table_name: str) -> bool:
    cur.execute(
        """
        SELECT COUNT(*) AS cnt
        FROM INFORMATION_SCHEMA.TABLES
        WHERE TABLE_SCHEMA = DATABASE()
          AND TABLE_NAME = %s
        """,
        (table_name,),
    )
    row = cur.fetchone()
    return bool(row and row.get("cnt"))


def _fetch_order_rows(from_date: Optional[str], to_date: Optional[str]) -> List[Dict[str, Any]]:
    where_sql, params = _where_by_date(from_date, to_date)
    db = MySQLClient()
    with db.cursor() as cur:
        cur.execute(
            f"""
            SELECT
                o.id,
                o.order_code,
                o.order_type,
                o.platform,
                o.shop_name,
                o.customer_name,
                o.customer_phone,
                o.order_status,
                o.packing_status,
                o.conveyor_status,
                o.shipping_status,
                o.note,
                o.created_at,
                o.updated_at,
                s.tracking_code,
                s.carrier_code,
                s.carrier_name,
                s.raw_last_status,
                COALESCE(v.video_count, 0) AS video_count,
                pv_last.file_path AS last_video_path
            FROM orders o
            LEFT JOIN shipments s ON s.id = (
                SELECT s2.id
                FROM shipments s2
                WHERE s2.order_id = o.id
                ORDER BY s2.updated_at DESC, s2.id DESC
                LIMIT 1
            )
            LEFT JOIN (
                SELECT
                    order_code,
                    COUNT(*) AS video_count,
                    MAX(id) AS last_video_id
                FROM packing_videos
                GROUP BY order_code
            ) v ON v.order_code = o.order_code
            LEFT JOIN packing_videos pv_last ON pv_last.id = v.last_video_id
            {where_sql}
            ORDER BY o.created_at DESC, o.id DESC
            """,
            params,
        )
        return list(cur.fetchall())


def _fetch_employee_rows(active_only: bool = False) -> List[Dict[str, Any]]:
    where_sql = "WHERE is_active = 1" if active_only else ""
    db = MySQLClient()
    with db.cursor() as cur:
        cur.execute(
            f"""
            SELECT
                employee_code,
                employee_name,
                department,
                phone,
                is_active,
                created_at,
                updated_at
            FROM employees
            {where_sql}
            ORDER BY is_active DESC, employee_code ASC
            """
        )
        return list(cur.fetchall())


def _format_small_package_lines(rows: Iterable[Dict[str, Any]]) -> Dict[str, str]:
    details: dict[str, list[str]] = defaultdict(list)
    for row in rows:
        order_code = _clean(row.get("order_code"))
        line_no = len(details[order_code]) + 1
        parts = [
            _clean(row.get("small_package_code")),
            f"Số lượng/Số lần quét: {int(row.get('scan_count') or 0)}",
        ]
        if row.get("scanners"):
            parts.append(f"Máy quét: {_clean(row.get('scanners'))}")
        if row.get("employees"):
            parts.append(f"NV: {_clean(row.get('employees'))}")
        if row.get("first_scanned_at"):
            parts.append(f"Lần đầu: {_clean(row.get('first_scanned_at'))}")
        if row.get("last_scanned_at"):
            parts.append(f"Lần cuối: {_clean(row.get('last_scanned_at'))}")
        details[order_code].append(f"{line_no}. " + " | ".join(parts))

    return {order_code: "\n".join(lines) for order_code, lines in details.items()}


def _fetch_small_package_summary(order_codes: Iterable[str]) -> Dict[str, str]:
    codes = [c for c in dict.fromkeys(order_codes) if c]
    if not codes:
        return {}

    placeholders = ", ".join(["%s"] * len(codes))
    db = MySQLClient()

    with db.cursor() as cur:
        if _table_exists(cur, "packing_small_packages"):
            cur.execute(
                f"""
                SELECT
                    p.order_code,
                    p.small_package_code,
                    COUNT(*) AS scan_count,
                    MIN(p.scanned_at) AS first_scanned_at,
                    MAX(p.scanned_at) AS last_scanned_at,
                    GROUP_CONCAT(DISTINCT p.scanner_id ORDER BY p.scanner_id SEPARATOR ', ') AS scanners,
                    GROUP_CONCAT(
                        DISTINCT TRIM(CONCAT(
                            COALESCE(p.employee_code, ''),
                            CASE WHEN e.employee_name IS NULL OR e.employee_name = '' THEN '' ELSE ' - ' END,
                            COALESCE(e.employee_name, '')
                        ))
                        ORDER BY p.employee_code SEPARATOR ', '
                    ) AS employees
                FROM packing_small_packages p
                LEFT JOIN employees e ON e.employee_code = p.employee_code
                WHERE p.order_code IN ({placeholders})
                  AND IFNULL(p.is_deleted, 0) = 0
                GROUP BY p.order_code, p.small_package_code
                ORDER BY p.order_code ASC, MIN(p.id) ASC
                """,
                codes,
            )
            summary = _format_small_package_lines(cur.fetchall())
            if summary:
                return summary

        cur.execute(
            f"""
            SELECT
                w.master_order_code AS order_code,
                w.child_order_code AS small_package_code,
                COUNT(*) AS scan_count,
                MIN(w.scanned_at) AS first_scanned_at,
                MAX(w.scanned_at) AS last_scanned_at,
                GROUP_CONCAT(DISTINCT w.scanner_id ORDER BY w.scanner_id SEPARATOR ', ') AS scanners,
                GROUP_CONCAT(
                    DISTINCT TRIM(CONCAT(
                        COALESCE(w.employee_code, ''),
                        CASE WHEN e.employee_name IS NULL OR e.employee_name = '' THEN '' ELSE ' - ' END,
                        COALESCE(e.employee_name, '')
                    ))
                    ORDER BY w.employee_code SEPARATOR ', '
                ) AS employees
            FROM wholesale_box_items w
            LEFT JOIN employees e ON e.employee_code = w.employee_code
            WHERE w.master_order_code IN ({placeholders})
              AND IFNULL(w.is_deleted, 0) = 0
            GROUP BY w.master_order_code, w.child_order_code
            ORDER BY w.master_order_code ASC, MIN(w.id) ASC
            """,
            codes,
        )
        return _format_small_package_lines(cur.fetchall())


def _is_wholesale(row: Dict[str, Any]) -> bool:
    value = _clean(row.get("order_type")).upper()
    return value in {"WHOLESALE", "SI", "SỈ", "DON_SI", "DONG_SI"}


def _apply_sheet_style(ws, headers: list[str], wrap_columns: set[int]) -> None:
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin = Side(style="thin", color="D9E2F3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        max_lines = 1
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(
                horizontal="left",
                vertical="top",
                wrap_text=cell.column in wrap_columns,
            )
            if isinstance(cell.value, str):
                max_lines = max(max_lines, cell.value.count("\n") + 1)
        ws.row_dimensions[row[0].row].height = min(18 * max_lines, 120)

    for col_idx, header in enumerate(headers, start=1):
        letter = get_column_letter(col_idx)
        width = max(len(header) + 2, 12)
        for cell in ws[letter]:
            if cell.value:
                first_line = max(str(cell.value).splitlines() or [""], key=len)
                width = max(width, min(len(first_line) + 2, 55))
        ws.column_dimensions[letter].width = width


def _append_empty_row(ws, headers: list[str], message: str) -> None:
    row = [""] * len(headers)
    row[0] = 1
    row[1] = message
    ws.append(row)


def _fill_ecom_sheet(ws, rows: list[Dict[str, Any]]) -> None:
    ws.append(ECOM_HEADERS)
    if not rows:
        _append_empty_row(ws, ECOM_HEADERS, "Không có dữ liệu Ecom")
    for idx, row in enumerate(rows, start=1):
        shop = " / ".join(part for part in [_clean(row.get("platform")), _clean(row.get("shop_name"))] if part)
        ws.append(
            [
                idx,
                _clean(row.get("order_code")),
                shop,
                _clean(row.get("customer_name")),
                _clean(row.get("customer_phone")),
                _clean(row.get("order_status")),
                _clean(row.get("packing_status")),
                _clean(row.get("conveyor_status")),
                _clean(row.get("shipping_status") or row.get("raw_last_status")),
                _clean(row.get("tracking_code")),
                _clean(row.get("carrier_name") or row.get("carrier_code")),
                row.get("video_count") or 0,
                _clean(row.get("last_video_path")),
                _clean(row.get("note")),
                _clean(row.get("created_at")),
                _clean(row.get("updated_at")),
            ]
        )
    _apply_sheet_style(ws, ECOM_HEADERS, {13, 14})


def _fill_employee_sheet(ws, rows: list[Dict[str, Any]]) -> None:
    ws.append(EMPLOYEE_HEADERS)
    if not rows:
        _append_empty_row(ws, EMPLOYEE_HEADERS, "Không có dữ liệu nhân sự")
    for idx, row in enumerate(rows, start=1):
        ws.append(
            [
                idx,
                _clean(row.get("employee_code")),
                _clean(row.get("employee_name")),
                _clean(row.get("department")),
                _clean(row.get("phone")),
                "active" if int(row.get("is_active") or 0) == 1 else "inactive",
                _clean(row.get("created_at")),
                _clean(row.get("updated_at")),
            ]
        )
    _apply_sheet_style(ws, EMPLOYEE_HEADERS, set())


def _fill_wholesale_sheet(ws, rows: list[Dict[str, Any]], small_package_details: Dict[str, str]) -> None:
    ws.append(WHOLESALE_HEADERS)
    if not rows:
        _append_empty_row(ws, WHOLESALE_HEADERS, "Không có dữ liệu Sỉ")
    for idx, row in enumerate(rows, start=1):
        order_code = _clean(row.get("order_code"))
        ws.append(
            [
                idx,
                order_code,
                _clean(row.get("customer_name")),
                _clean(row.get("customer_phone")),
                _clean(row.get("order_status")),
                _clean(row.get("packing_status")),
                small_package_details.get(order_code, ""),
                _clean(row.get("shipping_status") or row.get("raw_last_status")),
                _clean(row.get("tracking_code")),
                _clean(row.get("carrier_name") or row.get("carrier_code")),
                row.get("video_count") or 0,
                _clean(row.get("last_video_path")),
                _clean(row.get("note")),
                _clean(row.get("created_at")),
                _clean(row.get("updated_at")),
            ]
        )
    _apply_sheet_style(ws, WHOLESALE_HEADERS, {7, 12, 13})


def export_order_status_excel(from_date: Optional[str] = None, to_date: Optional[str] = None) -> BytesIO:
    rows = _fetch_order_rows(from_date, to_date)
    ecom_rows = [row for row in rows if not _is_wholesale(row)]
    wholesale_rows = [row for row in rows if _is_wholesale(row)]
    small_package_details = _fetch_small_package_summary(row["order_code"] for row in wholesale_rows)

    wb = Workbook()
    ws_ecom = wb.active
    ws_ecom.title = "Ecom"
    ws_wholesale = wb.create_sheet("Sỉ")

    _fill_ecom_sheet(ws_ecom, ecom_rows)
    _fill_wholesale_sheet(ws_wholesale, wholesale_rows, small_package_details)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def export_employees_excel(active_only: bool = False) -> BytesIO:
    rows = _fetch_employee_rows(active_only=active_only)
    wb = Workbook()
    ws = wb.active
    ws.title = "Nhan su"
    _fill_employee_sheet(ws, rows)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output
