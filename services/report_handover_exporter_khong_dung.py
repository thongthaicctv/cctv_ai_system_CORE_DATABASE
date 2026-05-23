# services/report_handover_exporter.py
# -*- coding: utf-8 -*-

import os
import sqlite3
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter


def _parse_date_input(text):
    """
    Nhập dd/mm/yyyy -> trả về yyyy-mm-dd
    """
    text = str(text or "").strip()
    if not text:
        return ""

    try:
        return datetime.strptime(text, "%d/%m/%Y").strftime("%Y-%m-%d")
    except Exception:
        return ""


def export_handover_report(
    db_path="db/packing.db",
    output_dir="reports",
    from_date="",
    to_date="",
):
    os.makedirs(output_dir, exist_ok=True)

    from_date_sql = _parse_date_input(from_date)
    to_date_sql = _parse_date_input(to_date)

    now = datetime.now()
    output_file = os.path.join(
        output_dir,
        f"Bao_cao_giao_hang_{now.strftime('%Y%m%d_%H%M%S')}.xlsx"
    )

    conn = sqlite3.connect(db_path)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()

    where = []
    params = []

    if from_date_sql:
        where.append("date(created_at) >= ?")
        params.append(from_date_sql)

    if to_date_sql:
        where.append("date(created_at) <= ?")
        params.append(to_date_sql)

    where_sql = ""
    if where:
        where_sql = "WHERE " + " AND ".join(where)

    cur.execute(f"""
        SELECT
            id,
            delivery_order_code,
            packing_order_code,
            packing_session_id,
            packing_scanner_id,
            delivery_scanner_id,
            result,
            error_message,
            first_scan_time,
            second_scan_time,
            created_at
        FROM handover_sessions
        {where_sql}
        ORDER BY created_at DESC, id DESC
    """, params)

    rows = cur.fetchall()
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "Bao cao giao hang"

    headers = [
        "STT",
        "Thời gian",
        "Scanner giao",
        "Mã giao",
        "Mã thùng",
        "Scanner đóng",
        "Packing Session ID",
        "Kết quả",
        "Lỗi/Ghi chú",
        "Lần quét giao",
        "Lần quét thùng",
    ]

    ws.append(headers)

    for idx, row in enumerate(rows, start=1):
        result = row["result"] or ""

        if result == "success":
            result_text = "Giao thành công"
        elif result == "failed_wrong_box":
            result_text = "Sai thùng hàng"
        elif result == "failed_not_packed":
            result_text = "Đơn chưa được đóng"
        elif result == "failed_already_delivered":
            result_text = "Đơn đã giao trước đó"
        elif result == "failed_missing_box_code":
            result_text = "Chưa quét mã thùng"
        else:
            result_text = result

        ws.append([
            idx,
            row["created_at"] or "",
            row["delivery_scanner_id"] or "",
            row["delivery_order_code"] or "",
            row["packing_order_code"] or "",
            row["packing_scanner_id"] or "",
            row["packing_session_id"] or "",
            result_text,
            row["error_message"] or "",
            row["first_scan_time"] or "",
            row["second_scan_time"] or "",
        ])

    # Style
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(name="Times New Roman", size=13, bold=True, color="FFFFFF")
    normal_font = Font(name="Times New Roman", size=13)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left = Alignment(horizontal="left", vertical="center", wrap_text=True)

    thin = Side(style="thin", color="999999")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = border

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = normal_font
            cell.border = border
            cell.alignment = center

        # Cột lỗi/ghi chú căn trái
        row[8].alignment = left

        # Tô màu kết quả
        result_cell = row[7]
        if result_cell.value == "Giao thành công":
            result_cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif result_cell.value:
            result_cell.fill = PatternFill("solid", fgColor="FFC7CE")

    widths = [8, 22, 16, 22, 22, 16, 18, 24, 35, 22, 22]
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_file)
    return output_file